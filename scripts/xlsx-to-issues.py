#!/usr/bin/env python3
"""
📊 xlsx-to-issues.py — Converte Excel do TCC em issues GitHub

USO:
  python3 scripts/xlsx-to-issues.py                    # Gera issues em .github/issues/
  python3 scripts/xlsx-to-issues.py --create            # Cria issues via gh CLI
  python3 scripts/xlsx-to-issues.py --project 1         # Adiciona ao Project V2 #1
  python3 scripts/xlsx-to-issues.py --dry-run            # Preview sem criar

FONTES:
  dados/05-catalogo-incidentes-*.xlsx  → Issues de incidentes (1 por incidente)
  dados/07-matriz-prioridade-sla-*.xlsx → Labels de prioridade + milestones
  diagramas/08-matriz-raci-*.xlsx      → Issues de atividades do processo

SAÍDA:
  .github/issues/incidentes/INC-XXX-descricao.md
  .github/issues/processo/PROC-XX-atividade.md
  scripts/labels-generated.json
"""

import openpyxl
import os
import json
import re
import subprocess
import argparse
from pathlib import Path


# ─────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────

PRIORITY_MAP = {
    "P1": {"label": "P1-critico", "color": "b60205", "sla": "4h"},
    "P2": {"label": "P2-alto",    "color": "e11d48", "sla": "8h"},
    "P3": {"label": "P3-medio",   "color": "f97316", "sla": "24h"},
    "P4": {"label": "P4-baixo",   "color": "0ea5e9", "sla": "72h"},
}

CATEGORY_COLORS = {
    "Hardware":   "d73a4a",
    "Software":   "0075ca",
    "Rede":       "e4e669",
    "Periférico": "d4c5f9",
}

RACI_LABELS = {
    "R": "responsible",
    "A": "accountable",
    "C": "consulted",
    "I": "informed",
}


def slugify(text: str) -> str:
    """Converte texto em slug kebab-case."""
    text = text.lower().strip()
    text = re.sub(r'[àáâã]', 'a', text)
    text = re.sub(r'[éê]', 'e', text)
    text = re.sub(r'[íî]', 'i', text)
    text = re.sub(r'[óôõ]', 'o', text)
    text = re.sub(r'[úû]', 'u', text)
    text = re.sub(r'[ç]', 'c', text)
    text = re.sub(r'[^a-z0-9\s-]', '', text)
    text = re.sub(r'[\s]+', '-', text)
    text = re.sub(r'-+', '-', text)
    return text[:60].strip('-')


def find_xlsx(base_dir: str, pattern: str) -> str | None:
    """Encontra arquivo xlsx pelo padrão no nome."""
    for root, _, files in os.walk(base_dir):
        for f in files:
            if f.endswith('.xlsx') and pattern in f.lower():
                return os.path.join(root, f)
    return None


def calc_priority(impact: int, urgency: int) -> str:
    """Calcula prioridade pela matriz impacto × urgência."""
    score = impact * urgency
    if score >= 20:
        return "P1"
    elif score >= 12:
        return "P2"
    elif score >= 6:
        return "P3"
    else:
        return "P4"


# ─────────────────────────────────────────────
# PARSER: CATÁLOGO DE INCIDENTES
# ─────────────────────────────────────────────

def parse_catalogo(filepath: str) -> list[dict]:
    """Lê catálogo de incidentes e retorna lista de issues."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    issues = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue

        inc_id = str(row[0]).strip()
        categoria = str(row[1]).strip()
        subcategoria = str(row[2]).strip()
        descricao = str(row[3]).strip()
        sintomas = str(row[4]).strip() if row[4] else ""
        impacto = int(row[5]) if row[5] else 3
        urgencia = int(row[6]) if row[6] else 3
        tem_solucao = str(row[8]).strip().lower() if row[8] else "não"
        frequencia = str(row[9]).strip() if row[9] else "?"
        observacao = str(row[10]).strip() if row[10] else ""

        prioridade = calc_priority(impacto, urgencia)
        pri_info = PRIORITY_MAP[prioridade]

        labels = [
            f"cat-{slugify(categoria)}",
            f"sub-{slugify(subcategoria)}",
            pri_info["label"],
            "incidente",
        ]
        if tem_solucao in ("sim", "yes", "true"):
            labels.append("solucao-documentada")

        issue = {
            "id": inc_id,
            "title": f"🔧 {inc_id}: {descricao}",
            "labels": labels,
            "priority": prioridade,
            "category": categoria,
            "subcategory": subcategoria,
            "body": (
                f"## Incidente: {descricao}\n\n"
                f"| Campo | Valor |\n"
                f"|-------|-------|\n"
                f"| **ID** | `{inc_id}` |\n"
                f"| **Categoria** | {categoria} → {subcategoria} |\n"
                f"| **Impacto** | {impacto}/5 |\n"
                f"| **Urgência** | {urgencia}/5 |\n"
                f"| **Prioridade** | **{prioridade}** (SLA: {pri_info['sla']}) |\n"
                f"| **Frequência** | {frequencia}/mês |\n"
                f"| **Solução documentada?** | {'✅ Sim' if tem_solucao in ('sim', 'yes', 'true') else '❌ Não'} |\n\n"
                f"### Sintomas\n\n{sintomas}\n\n"
                + (f"### Observações\n\n{observacao}\n\n" if observacao else "")
                + "### Checklist de Resolução\n\n"
                "- [ ] Incidente reproduzido/confirmado\n"
                "- [ ] Causa raiz identificada\n"
                "- [ ] Solução aplicada e testada\n"
                "- [ ] Solução documentada na base de conhecimento\n"
                "- [ ] Tempo de resolução registrado\n"
            ),
            "projectFields": {
                "status": "Todo",
                "prioridade": prioridade,
                "categoria": categoria,
            },
        }
        issues.append(issue)

    return issues


# ─────────────────────────────────────────────
# PARSER: MATRIZ RACI
# ─────────────────────────────────────────────

def parse_raci(filepath: str) -> list[dict]:
    """Lê matriz RACI e retorna lista de issues de atividades."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Matriz RACI"]
    issues = []

    # Ler headers (papéis)
    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip().split('\n')[0] if cell.value else ""
        headers.append(val)

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=1):
        if not row[0]:
            continue

        atividade = str(row[0]).strip()
        observacao = str(row[-1]).strip() if row[-1] else ""

        # Mapear RACI
        raci_map = {}
        responsible = []
        for col_idx in range(1, len(headers) - 1):
            papel = headers[col_idx]
            valor = str(row[col_idx]).strip() if row[col_idx] else "-"
            if valor != "-":
                raci_map[papel] = valor
                if "R" in valor:
                    responsible.append(papel)

        # Gerar tabela RACI
        raci_table = "| Papel | Responsabilidade |\n|-------|------------------|\n"
        for papel, valor in raci_map.items():
            raci_desc = " / ".join([RACI_LABELS.get(v.strip(), v.strip()) for v in valor.split("/")])
            raci_table += f"| {papel} | **{valor}** ({raci_desc}) |\n"

        proc_id = f"PROC-{idx:02d}"
        issue = {
            "id": proc_id,
            "title": f"⚙️ {proc_id}: {atividade}",
            "labels": ["processo", "raci"],
            "body": (
                f"## Atividade: {atividade}\n\n"
                f"### Matriz RACI\n\n{raci_table}\n"
                + (f"### Observação\n\n{observacao}\n\n" if observacao else "\n")
                + f"### Responsável principal\n\n"
                f"{'→ '.join(responsible) if responsible else 'Não definido'}\n\n"
                "### Checklist de Implementação\n\n"
                "- [ ] Atividade mapeada no processo TO-BE\n"
                "- [ ] Responsáveis comunicados e treinados\n"
                "- [ ] Ferramenta/formulário configurado\n"
                "- [ ] Procedimento documentado na base de conhecimento\n"
                "- [ ] Testado em simulação\n"
            ),
            "projectFields": {
                "status": "Todo",
                "tipo": "processo",
            },
        }
        issues.append(issue)

    return issues


# ─────────────────────────────────────────────
# PARSER: SLA → LABELS + MILESTONES
# ─────────────────────────────────────────────

def parse_sla(filepath: str) -> dict:
    """Lê SLAs e gera configuração de labels e milestones."""
    wb = openpyxl.load_workbook(filepath)

    config = {"labels": [], "milestones": []}

    # Labels de prioridade
    for key, info in PRIORITY_MAP.items():
        config["labels"].append({
            "name": info["label"],
            "color": info["color"],
            "description": f"{key}: SLA resolução {info['sla']}",
        })

    # Ler SLAs detalhados
    if "SLAs InfoPro" in wb.sheetnames:
        ws = wb["SLAs InfoPro"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0]:
                config["milestones"].append({
                    "title": str(row[0]).strip(),
                    "description": str(row[1]).strip() if row[1] else "",
                    "sla_resposta": str(row[2]).strip() if row[2] else "",
                    "sla_resolucao": str(row[3]).strip() if row[3] else "",
                    "escalonamento": str(row[4]).strip() if row[4] else "",
                })

    return config


# ─────────────────────────────────────────────
# GERADOR: MARKDOWN COM FRONTMATTER
# ─────────────────────────────────────────────

def generate_issue_md(issue: dict) -> str:
    """Gera conteúdo markdown com frontmatter YAML."""
    labels_yaml = json.dumps(issue["labels"])
    fields_yaml = ""
    if "projectFields" in issue:
        fields_yaml = "\nprojectFields:\n"
        for k, v in issue["projectFields"].items():
            fields_yaml += f"  {k}: {v}\n"

    return (
        f"---\n"
        f"title: \"{issue['title']}\"\n"
        f"labels: {labels_yaml}\n"
        f"assignees: []\n"
        f"{fields_yaml}"
        f"---\n\n"
        f"{issue['body']}"
    )


# ─────────────────────────────────────────────
# WRITER: SALVAR ARQUIVOS
# ─────────────────────────────────────────────

def write_issues(issues: list[dict], output_dir: str, prefix: str):
    """Escreve issues como arquivos .md."""
    os.makedirs(output_dir, exist_ok=True)
    for issue in issues:
        slug = slugify(issue["title"].split(": ", 1)[-1] if ": " in issue["title"] else issue["title"])
        filename = f"{issue['id'].lower()}-{slug}.md"
        filepath = os.path.join(output_dir, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(generate_issue_md(issue))
        print(f"  ✅ {filepath}")


def write_labels(config: dict, output_path: str):
    """Escreve labels.json para setup automatizado."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    all_labels = config["labels"] + [
        {"name": "incidente",           "color": "d73a4a", "description": "Incidente do catálogo"},
        {"name": "processo",            "color": "0075ca", "description": "Atividade do processo TO-BE"},
        {"name": "raci",                "color": "bfd4f2", "description": "Atividade com RACI definido"},
        {"name": "solucao-documentada", "color": "0e8a16", "description": "Já possui solução na base de conhecimento"},
    ]

    # Labels de categoria
    for cat, color in CATEGORY_COLORS.items():
        all_labels.append({
            "name": f"cat-{slugify(cat)}",
            "color": color,
            "description": f"Categoria: {cat}",
        })

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_labels, f, ensure_ascii=False, indent=2)
    print(f"  ✅ {output_path} ({len(all_labels)} labels)")


# ─────────────────────────────────────────────
# CRIADOR: GITHUB CLI (--create)
# ─────────────────────────────────────────────

def create_issues_gh_cli(issues: list[dict], project_number: int | None = None, dry_run: bool = False):
    """Cria issues diretamente via GitHub CLI."""
    for issue in issues:
        labels = ",".join(issue["labels"])
        cmd = [
            "gh", "issue", "create",
            "--title", issue["title"],
            "--body", issue["body"],
            "--label", labels,
        ]

        if dry_run:
            print(f"  🔍 [DRY RUN] {issue['title']}")
            print(f"     Labels: {labels}")
            continue

        try:
            result = subprocess.run(cmd, capture_output=True, text=True, check=True)
            issue_url = result.stdout.strip()
            print(f"  ✅ Criada: {issue_url}")

            # Adicionar ao project se especificado
            if project_number and issue_url:
                subprocess.run(
                    ["gh", "project", "item-add", str(project_number),
                     "--owner", "@me", "--url", issue_url],
                    capture_output=True, text=True
                )
                print(f"     📋 Adicionada ao Project #{project_number}")

        except subprocess.CalledProcessError as e:
            print(f"  ❌ Erro: {e.stderr}")
        except FileNotFoundError:
            print("  ❌ gh CLI não encontrado. Instale: https://cli.github.com")
            return


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Converte Excel do TCC em issues GitHub"
    )
    parser.add_argument("--base", default=".", help="Diretório raiz do projeto")
    parser.add_argument("--create", action="store_true", help="Criar issues via gh CLI")
    parser.add_argument("--project", type=int, help="Número do Project V2 para vincular")
    parser.add_argument("--dry-run", action="store_true", help="Preview sem criar")
    parser.add_argument("--output", default=".github/issues", help="Diretório de saída")
    args = parser.parse_args()

    base = args.base
    output_base = os.path.join(base, args.output)

    print("📊 xlsx-to-issues — Conversor Excel → GitHub Issues")
    print("=" * 55)

    # ── 1. CATÁLOGO DE INCIDENTES ──
    catalogo_path = (
        find_xlsx(base, "catalogo-incidentes") or
        find_xlsx(base, "catalogo")
    )
    incidentes = []
    if catalogo_path:
        print(f"\n📁 Catálogo: {catalogo_path}")
        incidentes = parse_catalogo(catalogo_path)
        print(f"   {len(incidentes)} incidentes encontrados")
    else:
        print("\n⚠️  Catálogo de incidentes não encontrado")

    # ── 2. MATRIZ RACI ──
    raci_path = find_xlsx(base, "matriz-raci")
    atividades = []
    if raci_path:
        print(f"\n📁 RACI: {raci_path}")
        atividades = parse_raci(raci_path)
        print(f"   {len(atividades)} atividades encontradas")
    else:
        print("\n⚠️  Matriz RACI não encontrada")

    # ── 3. SLA → LABELS ──
    sla_path = find_xlsx(base, "prioridade-sla") or find_xlsx(base, "sla")
    sla_config = {"labels": [], "milestones": []}
    if sla_path:
        print(f"\n📁 SLA: {sla_path}")
        sla_config = parse_sla(sla_path)
        print(f"   {len(sla_config['labels'])} labels + {len(sla_config['milestones'])} milestones")
    else:
        print("\n⚠️  Matriz de SLA não encontrada")

    # ── GERAR OUTPUT ──
    total = len(incidentes) + len(atividades)

    if args.create or args.dry_run:
        print(f"\n{'🔍 DRY RUN' if args.dry_run else '🚀 CRIANDO'}: {total} issues via gh CLI")
        print("-" * 40)
        all_issues = incidentes + atividades
        create_issues_gh_cli(all_issues, args.project, args.dry_run)
    else:
        print(f"\n📝 Gerando {total} issue files em {output_base}/")
        print("-" * 40)

        if incidentes:
            print(f"\n🔧 Incidentes ({len(incidentes)}):")
            write_issues(incidentes, os.path.join(output_base, "incidentes"), "INC")

        if atividades:
            print(f"\n⚙️ Atividades do processo ({len(atividades)}):")
            write_issues(atividades, os.path.join(output_base, "processo"), "PROC")

        labels_path = os.path.join(base, "scripts", "labels-generated.json")
        print(f"\n🏷️ Labels:")
        write_labels(sla_config, labels_path)

    # ── RESUMO ──
    print(f"\n{'=' * 55}")
    print(f"✅ Concluído!")
    print(f"   📋 {len(incidentes)} issues de incidentes")
    print(f"   ⚙️ {len(atividades)} issues de atividades (RACI)")
    print(f"   🏷️ {len(sla_config['labels']) + len(CATEGORY_COLORS) + 4} labels configuradas")

    if not args.create:
        print(f"\n📌 Próximos passos:")
        print(f"   • Rodar com --create para criar issues via gh CLI")
        print(f"   • Ou push pro repo para o workflow seed-issues.yml criar automaticamente")
        print(f"   • Adicionar --project N para vincular ao Project V2")


if __name__ == "__main__":
    main()
